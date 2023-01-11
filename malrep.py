import pandas as pd
import openpyxl
import os
from global_defs import *

# ---------------------------------------------------------
#   MALREP - function library
# ---------------------------------------------------------

# add new malfuntion row to excel file
def add_malfunction_to_excel(input_file, new_entry):
    # verify file exists
    if not os.path.isfile(input_file):
        print(f"הקובץ לא נמצא-{input_file}")
        return ERR_ROW_NOT_APPENDED
    # load excel
    wb = openpyxl.load_workbook(filename=input_file)
    ws = wb[MALEX_MALFUNCTIONS_SHEET]
    # new row index
    row = ws.max_row + 1
    # add new row cells
    for col, entry in enumerate(new_entry, start=1):
        ws.cell(row=row, column=col, value=entry)
    # save excel file with new row
    wb.save(input_file)
    print(f"תקלה נרשמה {row}")
    # return the index of the new row
    return row


# Verify validity of selected input file
def verify_input_file(input_file):
    # file exists
    if not os.path.exists(input_file):
        print(f"File does not exist {input_file}")
        return False
    # file is xlsx format
    elif not input_file.split('.')[-1] == 'xlsx':
        print(f"File format isn't xlsx")
        return False
    return True


# Extract props of specific subsystem type
def get_prop_list_by_subsystem(props_list, subsystem):
    # verify subsystem is given as argument
    if not subsystem:
        # return all props regardless of subsystem
        return_list = [prop.name for prop in props_list]
        return_list.append(ELSE)

    return_list = []
    for prop in props_list:
        # match prop subsystem with argument subsystem
        if prop.subsystem == subsystem:
            if prop.name:
                return_list.append(prop.name)
    # add 'else' option
    return_list.append(ELSE)
    return return_list


# Validate entry from excel
def validate_entry(list, entry):
    if entry in list:
        return list
    if not entry:
        return list
    list.append(entry)
    return list


# Extract all lists from excel 'listings' sheet
def get_all_systems_lists(input_file):
    if not os.path.isfile(input_file):
        # if no file was found, return dummy lists
        print(f"הקובץ לא נמצא-{input_file}")
        return [''],[''],[''],['']

    systems_list = []
    subsystems_list = []
    malfunctions_list = []
    malfunctions_status_list = []
    # load excel and iterate over rows
    lists_sheet = pd.read_excel(input_file, sheet_name=MALEX_LISTS_SHEET, keep_default_na=False, skiprows=[0])
    for index, row in lists_sheet.iterrows():
        systems_list    = validate_entry(systems_list ,str(row[COL_SYSTEM]))
        subsystems_list  = validate_entry(subsystems_list ,str(row[COL_SUBSYSTEM]))
        malfunctions_list = validate_entry(malfunctions_list ,str(row[COL_MALFUNCTION]))
        malfunctions_status_list = validate_entry(malfunctions_status_list ,str(row[COL_MALFUNCTION_STAUTS]))
    # add 'else' option
    systems_list.append(ELSE)
    subsystems_list.append(ELSE)
    malfunctions_list.append(ELSE)
    malfunctions_status_list.append(ELSE)
    return systems_list, subsystems_list, malfunctions_list, malfunctions_status_list


# load props from excel
def load_malfunctions_excel(input_file):
    if not os.path.isfile(input_file):
        # if file not found, return dummy prop
        print(f"הקובץ לא נמצא-{input_file}")
        return [Prop(   name='',
                        system= '',
                        subsystem= '',
                        description= '')]

    props_list = []
    props_sheet = pd.read_excel(open(input_file, 'rb'), sheet_name=MALEX_PROPS_SHEET, keep_default_na=False)
    for index, row in props_sheet.iterrows():
        system = str(row[COL_SYSTEM_IDX])
        subsystem = str(row[COL_SUBSYSTEM_IDX])
        name = str(row[COL_NAME_IDX])
        description = str(row[COL_DESCRIPTION_IDX])
        # create new prop struct
        prop = Prop(name=name,
                    system= system,
                    subsystem= subsystem,
                    description= description)

        props_list.append(prop)

    return props_list

